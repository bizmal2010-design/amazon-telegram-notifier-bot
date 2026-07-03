import asyncio
import random
import re
import sqlite3
import aiosqlite
import openpyxl
import os
import sys
from patchright.async_api import async_playwright
import tkinter as tk 
from tkinter import messagebox
import pandas as pd
import subprocess

def get_base_path():
    """يعيد المسار الصحيح سواء كان البرنامج مجمّع (EXE) أو بيثون عادي"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()
DB_PATH = os.path.join(BASE_PATH, "data.db")
EXCEL_PATH = os.path.join(BASE_PATH, "amazon_products.xlsx")
BROWSERS_PATH = os.path.join(BASE_PATH, "browsers")
BROWSERS_PATH = os.path.join(BASE_PATH, "browsers")
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = BROWSERS_PATH
def init_db():
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id          INTEGER PRIMARY KEY,
                title       TEXT,
                price       REAL,
                rating      TEXT,
                rating_count TEXT,
                image_url   TEXT,
                link        TEXT,
                asin        TEXT UNIQUE
            )
        ''')
        conn.commit()
    except sqlite3.OperationalError as e:
        print(f"[DB ERROR] قاعدة البيانات مشغولة: {e}")
    finally:
        conn.close()


# ===== دالة الكشط =====
async def scrape_amazon():
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = BROWSERS_PATH

    if not os.path.exists(BROWSERS_PATH):
        messagebox.showerror(
            "خطأ",
            f"لم يتم العثور على مجلد المتصفحات:\n{BROWSERS_PATH}\n\n"
            "تأكد من وجود مجلد 'browsers' بجانب ملف البرنامج."
        )
        return 0

    # ✅ الإصلاح 1: حذف البيانات القديمة قبل بدء الكشط
    try:
        async with aiosqlite.connect(DB_PATH) as conn:
            await conn.execute("DELETE FROM products")
            await conn.commit()
        print("[INFO] تم مسح البيانات القديمة من قاعدة البيانات.")
    except Exception as e:
        print(f"[ERROR] فشل مسح البيانات القديمة: {e}")

    # ✅ الإصلاح 1: حذف ملف Excel القديم إن وجد
    if os.path.exists(EXCEL_PATH):
        try:
            os.remove(EXCEL_PATH)
            print("[INFO] تم حذف ملف Excel القديم.")
        except Exception as e:
            print(f"[ERROR] فشل حذف ملف Excel القديم: {e}")

    all_products = []

    async with async_playwright() as p:
        print("[LOADING] تشغيل المتصفح في الوضع الصامت...")
        browser = await p.chromium.launch(
            headless=True,
            channel="chrome",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/137.0.0.0 Safari/537.36"
            ),
            extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="America/New_York",
            color_scheme="light"
        )

        page = await context.new_page()
        search_url = "https://www.amazon.com/s?k=laptops"
        print(f"[INFO] الاتصال بـ: {search_url}")
        await page.goto(
            search_url,
            wait_until="domcontentloaded",
            timeout=60000
        )
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            """)
        await page.wait_for_selector(
            "div[data-component-type='s-search-result']",
            timeout=30000
        )

        print(await page.title())
        print(page.url)

        await page.screenshot(path="amazon.png", full_page=True)
        page_num = 1
        for _ in range(3):
            for i in range(1, 6):
                await page.evaluate(
                    f"window.scrollTo(0, (document.body.scrollHeight * {i}) / 5);"
                )
                await page.wait_for_timeout(random.randint(800, 1500))
            print(f"[INFO] فحص بطاقات المنتجات في الصفحة {page_num}...")
            products = await page.query_selector_all(
                "div[data-component-type='s-search-result']"
            )
            print(f"[INFO] وُجد {len(products)} منتج في الصفحة {page_num}.")

            page_items = 0

            for index, product in enumerate(products, 1):
                title_elem = await product.query_selector("h2 span")
                title = (await title_elem.inner_text()).strip() if title_elem else "Unknown"

                price = None
                price_elem = await product.query_selector("span.a-price span.a-offscreen")
                if not price_elem:
                    price_elem = await product.query_selector(
                        "div[data-cy='secondary-offer-recipe'] span.a-color-base"
                    )
                if price_elem:
                    try:
                        price_text = (await price_elem.inner_text()).strip()
                        price_text = re.sub(r'[^\d.]', '', price_text)  # يحذف كل شيء ما عدا الأرقام والنقطة
                        price = float(price_text) if price_text else None
                    except Exception as e:
                        print(f"[ERROR] فشل استخراج السعر للمنتج {index}: {e}")

                rating = "No Rating"
                rating_elem = await product.query_selector("span.a-icon-alt")
                if rating_elem:
                    rating = (await rating_elem.inner_text()).replace("out of 5 stars", "").strip()

                rating_count = "0"
                rating_man = await product.query_selector("a[aria-label*='ratings']")
                if rating_man:
                    attr = await rating_man.get_attribute("aria-label")
                    if attr:
                        rating_count = attr.strip().replace(" ratings", "")

                image_url = "N/A"
                image_elem = await product.query_selector("img.s-image")
                if image_elem:
                    image_url = await image_elem.get_attribute("src") or "N/A"

                asin = (await product.get_attribute("data-asin") or "").strip() or "N/A"
                link = f"https://www.amazon.com/dp/{asin}" if asin != "N/A" else "N/A"

                if title == "Unknown" or price is None or price < 100.0:
                    print(f"[SKIP] تم تخطي المنتج {index} (عنوان أو سعر مفقود).")
                    continue

                all_products.append(
                    [title, price, rating, rating_count, image_url, link, asin]
                )
                page_items += 1

                if index % 5 == 0 or index == len(products):
                    print(f"   ⚡ تم كشط {index}/{len(products)} عنصر...")

            print(f"[SUCCESS] تم استخراج {page_items} منتج من الصفحة {page_num}.")

            next_btn = await page.query_selector("a.s-pagination-next")
            if next_btn:
                # احذف الـ popup إن وجد
                try:
                    await page.evaluate("""
                        const modal = document.querySelector('div#redir-modal');
                        if (modal) modal.remove();
                        const overlay = document.querySelector('div#redir-overlay');
                        if (overlay) overlay.remove();
                    """)
                except:
                    pass

                # انتقل بالرابط مباشرة بدل click
                next_url = await next_btn.get_attribute("href")
                if next_url:
                    if not next_url.startswith("http"):
                        next_url = "https://www.amazon.com" + next_url
                    print("[INFO] Moving to next page...")
                    await page.goto(next_url)
                    await page.goto(
                        next_url,
                        wait_until="networkidle"
                    )
                    await page.wait_for_selector(
                        "div[data-component-type='s-search-result']"
                    )
                    await page.wait_for_timeout(random.randint(2000, 4000))
                    page_num += 1
            else:
                print("[INFO] No more pages.")
                break

        print("\n[INFO] إغلاق المتصفح...")
        await browser.close()

    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.executemany(
            """
            INSERT OR REPLACE INTO products
            (title, price, rating, rating_count, image_url, link, asin)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            all_products
        )
        await conn.commit()

    return len(all_products)


# ✅ الإصلاح 3: التحقق من وجود بيانات في DB بشكل صحيح
def check_scraping_status():
    print(f"[DEBUG] DB_PATH = {DB_PATH}")
    print(f"[DEBUG] file exists = {os.path.exists(DB_PATH)}")
    
    if not os.path.exists(DB_PATH):
        messagebox.showwarning("تنبيه", "يجب تشغيل عملية الكشط أولاً!")
        return False

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM products")
        count = cursor.fetchone()[0]
        conn.close()
        print(f"[DEBUG] count = {count}")
        
        if count == 0:
            messagebox.showwarning("تنبيه", "يجب تشغيل عملية الكشط أولاً!")
            return False
        return True
    except Exception as e:
        print(f"[DEBUG] ERROR: {e}")
        messagebox.showerror("خطأ", str(e))
        return False

# ===== أزرار الواجهة =====
def run_scraper():
    try:
        btn.config(state="disabled", text="جاري الكشط... يرجى الانتظار")
        root.update()
        count = asyncio.run(scrape_amazon())
        if count > 0:
            messagebox.showinfo("نجاح", f"تم كشط {count} منتج وحفظها في قاعدة البيانات!")
    except Exception as e:
        messagebox.showerror("خطأ", f"حدث خطأ:\n{str(e)}")
    finally:
        btn.config(state="normal", text="ابدأ عملية الكشط الآن")


def export_to_excel():
    if not check_scraping_status():
        return
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query("SELECT * FROM products", conn)
        conn.close()

        if df.empty:
            messagebox.showwarning("تنبيه", "قاعدة البيانات فارغة!\nابدأ عملية الكشط أولاً.")
            return

        df.to_excel(EXCEL_PATH, index=False, engine='openpyxl')

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            for col_idx in [6, 7]:
                cell = ws.cell(row=r, column=col_idx)
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        wb.save(EXCEL_PATH)

        messagebox.showinfo("نجاح", f"تم التصدير بنجاح!\n{EXCEL_PATH}")

    except Exception as e:
        messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير:\n{str(e)}")


def open_db_browser():
    if not check_scraping_status():
        return
    db_browser_paths = [
        r"C:\Program Files\DB Browser for SQLite\DB Browser for SQLite.exe",
        r"C:\Program Files (x86)\DB Browser for SQLite\DB Browser for SQLite.exe",
        r"C:\Program Files\DB Browser for SQLCipher\DB Browser for SQLCipher.exe",
        r"C:\Program Files (x86)\DB Browser for SQLCipher\DB Browser for SQLCipher.exe",
    ]

    exe_path = None
    for path in db_browser_paths:
        if os.path.exists(path):
            exe_path = path
            break

    if exe_path is None:
        messagebox.showerror("خطأ", "لم يتم العثور على DB Browser!\nتأكد من تثبيته على جهازك.")
        return

    if not os.path.exists(DB_PATH):
        messagebox.showwarning("تنبيه", "قاعدة البيانات غير موجودة!\nابدأ عملية الكشط أولاً.")
        return

    subprocess.Popen([exe_path, DB_PATH])


def open_excel_file():
    # ✅ الإصلاح 2: التحقق من وجود ملف Excel أولاً قبل أي شيء آخر
    if not os.path.exists(EXCEL_PATH):
        messagebox.showwarning(
            "تنبيه",
            "لم يتم تصدير ملف Excel بعد!\nقم بالضغط على 'تصدير النتائج إلى Excel' أولاً."
        )
        return

    if not check_scraping_status():
        return

    if sys.platform == "win32":
        os.startfile(EXCEL_PATH)
    elif sys.platform == "darwin":
        os.system(f'open "{EXCEL_PATH}"')
    else:
        os.system(f'xdg-open "{EXCEL_PATH}"')


# ===== الواجهة الرسومية =====
root = tk.Tk()
root.title("برنامج كشط البيانات الذكي")
root.geometry("800x460")
root.resizable(False, False)

tk.Label(
    root, text="نظام استخراج البيانات المطور",
    font=("Arial", 14, "bold")
).pack(pady=15)

btn = tk.Button(
    root, text="ابدأ عملية الكشط الآن",
    command=run_scraper,
    bg="green", fg="white", font=("Arial", 12), width=28
)
btn.pack(pady=5)

tk.Button(
    root, text="تصدير النتائج إلى Excel",
    command=export_to_excel,
    bg="blue", fg="white", font=("Arial", 12), width=28
).pack(pady=5)

tk.Button(
    root, text="فتح قاعدة البيانات",
    command=open_db_browser,
    bg="purple", fg="white", width=28
).pack(pady=4)

tk.Button(
    root, text="فتح ملف الإكسيل",
    command=open_excel_file,
    bg="orange", fg="black", width=28
).pack(pady=4)

if __name__ == "__main__":
    print(f"[DEBUG] BASE_PATH = {BASE_PATH}")
    print(f"[DEBUG] DB_PATH = {DB_PATH}")
    print(f"[DEBUG] db exists = {os.path.exists(DB_PATH)}")
    init_db()
    root.mainloop()
    
